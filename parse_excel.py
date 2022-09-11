import openpyxl
import re

MANUAL_MAPPING_XLSX = "/home/shovalg@staff.technion.ac.il/PycharmProjects/ObjectMappingNN/real_img_manual_mapping.xlsx"
PREFIX = ',"obj_params":{'
SUFFIX = '}'
X_PAT = re.compile(r"X_: (\d+)")
Y_PAT = re.compile(r"Y_: (\d+)")


def parse_params():
    wb = openpyxl.open(MANUAL_MAPPING_XLSX, data_only=True)
    sheet = wb.active
    row_idx = 0
    for row in sheet.iter_rows():
        if row_idx == 0:
            row_idx += 1
            continue
        row_idx += 1
        for idx in range(1, len(row), 2):
            part = row[idx:idx+2]
            if all(cell.value for cell in part):
                params = part[1].value
                category = part[0].value
                x_vals = X_PAT.findall(params)
                y_vals = Y_PAT.findall(params)
                if category.endswith('ball'):
                    ball_json_format(x_vals, y_vals, sheet, row_idx, idx+2)
                elif category == 'cart':
                    cart_json_format(x_vals, y_vals, sheet, row_idx, idx+2)
                else:
                    corner_json_format(x_vals, y_vals, sheet, row_idx, idx+2)
    wb.save(MANUAL_MAPPING_XLSX)


def cart_json_format(x_vals, y_vals, sheet, row_idx, col_idx):
    xs = [int(x) for x in x_vals]
    ys = [int(y) for y in y_vals]
    final_str = PREFIX
    for i in range(4):
        final_str += f'"X_corner_{i + 1}":{xs[i]},"Y_corner_{i + 1}":{ys[i]},'
    final_str += f'"X_center":{xs[4]},"Y_center":{ys[4]},"X_center_2":{xs[5]},"Y_center_2":{ys[5]},"radius":{xs[6] - xs[5]}{SUFFIX}'
    print(final_str)
    sheet.cell(row_idx, col_idx, final_str)


def corner_json_format(x_vals, y_vals, sheet, row_idx, col_idx):
    xs = [int(x) for x in x_vals]
    ys = [int(y) for y in y_vals]
    final_str = PREFIX
    for i in range(len(xs)):
        final_str += f'"X_corner_{i + 1}":{xs[i]},"Y_corner_{i + 1}":{ys[i]},'
    print(final_str[:-1] + SUFFIX)
    sheet.cell(row_idx, col_idx, final_str[:-1] + SUFFIX)
    

def ball_json_format(x_vals, y_vals, sheet, row_idx, col_idx):
    xc = int(x_vals[0])
    yc = int(y_vals[0])
    r = int(x_vals[1]) - xc
    final_str = f'{PREFIX}"X_center":{xc},"Y_center":{yc},"radius":{r}{SUFFIX}'
    print(final_str)
    sheet.cell(row_idx, col_idx, final_str)


if __name__ == '__main__':
    parse_params()
